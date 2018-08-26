from openpyxl import load_workbook, Workbook
from os import listdir
import os
from os.path import isfile, join,isdir
from DialogLib import *
from tClasses import sObj
from docx import Document
import re
from ckWord import docTodocx

def saveSectionData(sList):
    wb = Workbook()
    ws = wb.create_sheet("Sections",0)

    #Use Labels
    ws["A1"] =  "ID"
    ws["B1"] =  "Section Number"
    ws["C1"] =  "Section Name"
    ws["D1"] = "Pages"
    ws["E1"] =  "Related"
    ws["F1"] = "Required?"
    ws["G1"] = "Division"
    
    i = 2
    
    for s in sList:
        ws["A" + str(i)] =  s.sID
        ws["C" + str(i)]=  s.sNum
        ws["B" + str(i)] =  s.sName

        #Related
        try:
            #Print out List
            tString = ""
            for r in s.related:
                tString = tString + r + ","
            tString = tString[:-1]
            
            ws["E" + str(i)] =  tString
        except:
            pass

        try:
            if int(s.division) == 1:
                ws["F" + str(i)] =  1
            else:
                ws["F" + str(i)] =  0
        except:
            ws["F" + str(i)] =  1
            
        ws["G" + str(i)] =  s.division
        i = i + 1

    wb.save("Sections.xlsx")

def getSectionList():
    #Pick Directory for section
    mypath = pickDirectory()

    #List of sections
    sList =[]

    #Directory List
    directories = [d for d in listdir(mypath) if isdir(join(mypath,d))]

    sList.extend(getSection(mypath))
    
    #Directory Loop to get section in each folder (one level)
    for d in directories:
        sList.extend(getSection(join(mypath,d)))

    #Section List Processing
    i = 1

    #ID Assignment
    for s in sList:
        s.sID = i
        i = i + 1

    return sList
    
def getSection(mypath):
    #Temp List
    tList = []
    
    #File Loop
    files = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    
    for fs in files:
        
        #Start section object
        tObj = sObj()

        if fs.find("Appendix") != -1:
            tNum = "Appendix"
            tName = fs
            tDiv = "A"
        else:
            #Detect Name Features
            i = fs.find("-")

            #Split them up
            tNum = fs[:i-1]
            tName = fs[i+1:]

        #Remove end extension
        tName = tName.replace(".docx","")
        tName = tName.replace(".doc","")

        if tNum != "Appendix":  
            #Remove whitespace
            tNum = tNum.replace(" ","")

            #Division Number
            tDiv = str(tNum[0:2])
        
            #Add Spaces between numbers
            tNum = tNum[0:2]+" "+tNum[2:4]+" "+tNum[4:]

            #Get Related Sections if any
            #Check if doc
            nFile = docTodocx(str(fs),mypath)
            tObj.related = searchSection(join(mypath,nFile))

        #Assign object
        tObj.sNum = tNum
        tObj.sName = tName
        tObj.division = tDiv
        
        #Add to list
        tList.append(tObj)

    return tList

def searchSection(dpath):
    document = Document(dpath)
    content = document.paragraphs
    
    tRelated = []

    #Compiling pattern
    pattern = re.compile(r'(Section?\s?)(\d\d\s?\d\d\s?\d\d)')

    for text in content:
        matches = pattern.finditer(text.text)
        for match in matches:
            if len(match.group(2)) >= 8:
                tRelated.append(match.group(2))
            elif len(match.group(2)) < 8:
                tString = str(match.group(2))
                tRelated.append(tString[0:2]+" "+tString[2:4]+" "+tString[4:])
            else:
                pass

    re.purge()

    return tRelated
