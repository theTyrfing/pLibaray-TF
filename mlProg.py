from openpyxl import load_workbook, Workbook
from os import listdir
import os
from os.path import isfile, join,isdir
from DialogLib import *
from tClasses import *
from nltk import tokenize

def saveLite(iList):
    wb = Workbook()
    ws = wb.create_sheet("Bid_Data",0)

    i = 1
    
    for item in iList:
        if item.nID == 1:
            ws.cell(row = i, column = 1).value = item.iDecp
            i = i + 1

    wb.save("bidData.xlsx")

def mlData():
    mypath = pickFile("xlsx")
    wb = load_workbook(mypath)
    ws = wb["Bid_Data"]

    dList = []
    pList = []
    i = 1

    while True:
        if ws.cell(row = i, column = 1).value != None:
            dList.append(ws.cell(row = i, column = 1).value)
            i = i + 1
        else:
            break

    i = 1
    
    for data in dList:
        pdata = tokenize.sent_tokenize(data)
        if isinstance(pdata, list):
            for p in pdata:
                pList.append(chData(p,i))
        else:
            pList.append(chData(pdata[0],i))
        i = i + 1

    return pList

def mlSave():
    print("Preparing script, please answer the following for the machine....")
    pList = mlData()
    wb = Workbook()
    ws = wb.create_sheet("Bid_Data_X",0)
    i = 1

    for p in pList:
        ws.cell(row = i, column = 1).value = p.sData
        ws.cell(row = i, column = 2).value = p.sNum
        i = i + 1

    wb.save("bidDataProcessed.xlsx")

def chData(sentence, num):
    print("---------------")
    print("Entry [" + str(num) + "]")
    print("---------------")
    print(sentence)
    print("---------------")
    sAns = input("Input Section Number:")

    tObj = dObj()
    tObj.sData = str(sentence)
    tObj.sNum = str(sAns)

    return tObj
