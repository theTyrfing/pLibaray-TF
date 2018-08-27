from openpyxl import load_workbook, Workbook
from os import listdir
import os
from os.path import isfile, join,isdir
from DialogLib import *
from sProg import *
from tClasses import itemObj, sObj, cObj
import re
from time import sleep
from configparser import ConfigParser

unitPath =""
sectionPath = ""

def gatherBids():
    #Item Object Array
    iList = []
    
    #Pick a directory to get information
    pathString = pickDirectory()

    file_names = [os.path.join(pathString, f) 
        for f in os.listdir(pathString) 
        if f.endswith("xlsx")]
        
    for f in file_names:
        print( "Processing file: "+f)
        if str(f).find(".xlsx") != -1:
            iList.extend(compileItems(str(f)))
    return iList
    
def compileItems(fname):
    wb = load_workbook(fname,data_only=True)
    try:
        ws = wb["Detailed"]
    except:
        ws = wb.worksheets[0]        

    t2List = []    

    nGame = findRows(ws)
    endCol = findCol(ws,nGame[0])

    cList = getContractors(wb,endCol)
    sProj, sDate, sLoc = getInfo(ws)

    print("Compile items")
    
    for i in nGame:
        t2List.extend(getItem(sProj, sDate, sLoc, ws,cList,i,endCol))
        
    return t2List
    
def findRows(ws):
    nGame = []
    
    x = 2
    y = 15
    coBlank = 0
    
    while True:
        if ws.cell(row = y, column = x).value == None:
            coBlank = coBlank + 1
            
        else:
            coBlank = 0
            if ckNum(str(ws.cell(row = y, column = x).value))==True:
                nGame.append(y)

        if coBlank == 20:
            break

        y = y + 1
    return nGame

def findCol(ws,row):
    x = 2
    y = row
    coBlank = 0

    while True:
        if ws.cell(row = y, column = x).value == None:
            coBlank = coBlank + 1
            
        else:
            coBlank = 0

        if coBlank == 5:
            break

        x = x + 1

    x = x - 5
    
    return x

def getContractors(wb,col):
    try:
        ws_sum = wb["Summary ALL"]
    except:
        ws_sum = wb.worksheets[0]
        
    contractor = []
    
    for row in ws_sum.iter_rows(min_row=14, min_col= 2, max_col=col, max_row=15):
        for cell in row:
            if str(cell.value).find("Base BID") != -1:
                pass
            elif str(cell.value).find("Base Bid") != -1:
                pass
            elif str(cell.value).find("BASE BID") != -1:
                pass
            elif str(cell.value).find("AVERAGE BID COST") != -1:
                pass
            elif str(cell.value).find("Unit Price") != -1:
                pass
            elif (str(cell.value).upper()).find("BID SUMMARY") != -1:
                pass
            elif ckNum(str(cell.value)):
                pass
            elif re.fullmatch(r'\s+',str(cell.value)) != None:
                pass
            elif cell.value == None:
                pass
            else:
                contractor.append(str(cell.value))
    return contractor

def getInfo(ws):
    sProj = ""
    sDate = ""
    sLoc = ""
    altLoc = ""
    addPattern = r'(\d+\s(St\.)?\s?([a-zA-Z]+|\s)+[a-zA-Z]+\.?\s?[West|East|North|South|W|E|N|S]?)'
    projPattern = r'(Project\s*:\s*)(.*)'
    datePattern = r'(Date\s*:\s*)(.*)'
    altPattern = r'((MTCC|TSCC|YRCC|YCC|ACC|BCC|YNCC|YRSCC|YRCECC|YRVLCC|SCC|SSCC|PSCC|PCC|OCC|DCC|DSCC|GSCC|HCC|HSCC)\s?(#|No|No\.)?\s?\d+)'
    for row in ws.iter_rows(min_row=9, min_col= 2, max_col=6, max_row=12):
        for cell in row:
            iLoc = fdString(addPattern,str(cell.value),0)
            iProject = fdString(projPattern,str(cell.value),2)
            iDate = fdString(datePattern,str(cell.value),2)
            ialtLoc = fdString(altPattern,str(cell.value),0)
            if iProject != None:
                sProj = iProject
            elif iDate != None:
                sDate = iDate
            elif iLoc != None:
                sLoc = iLoc
            elif ialtLoc != None:
                altLoc = ialtLoc
            else:
                pass
            
    if sLoc == "" and altLoc != "":
        sLoc = altLoc
        
    return sProj, sDate, sLoc

def fdString(pattern, string, num):
    match = re.search(pattern,string)
    if match != None:
        return match.group(num)
    else:
        return None

def getItem(sProj, sDate, sLoc, ws,cList,row,col):
    tList = []
    i = 0
    numBool = False
    LS = False
    unitYes = False
    quantity = 1
    unitValue = ""
    iID = 1
    for row in ws.iter_rows(min_row=row, min_col= 3, max_col=col, max_row=row):
        for cell in row:
            if str(cell.value).find("N/A") != -1:
                pass
            elif str(cell.value).find("Lump Sum") != -1 or str(cell.value) == "LS":
                unitValue = "LS"
                LS = True
            elif str(cell.value) == "Allowance" or str(cell.value) == "Allow.":
                unitValue = "Allow."
                LS = True
            elif str(cell.value) == "m":
                unitValue = "m"
                unitYes = True
            elif str(cell.value) == "m2" or str(cell.value) == "sq.m":
                unitValue= "sq.m"
                unitYes = True
            elif str(cell.value) == "EA":
                unitValue = "EA"
                unitYes = True
            elif len(str(cell.value)) > 5 and ckNum(str(cell.value)) == False:
                dString = str(cell.value).replace("\n"," ")
                #dString = dString.replace("\t","")
            elif ckNum(str(cell.value)) == True:
                if i < len(cList):
                    if LS == False and numBool == True:
                        numBool = False
                        pass
                    elif LS == True and cell.value == 1:
                        pass
                    else:
                        if unitYes == True:
                            quantity = cell.value
                            unitYes = False
                        else:
                            tObj = itemObj()
                            tObj.project = sProj + "-" + sLoc
                            tObj.date = sDate
                            tObj.iDecp = dString
                            tObj.unit = unitValue
                            tObj.qu = quantity
                            tObj.unitRate = cell.value
                            tObj.contractor = cList[i]
                            tObj.nID = iID
                            i = i + 1
                            tList.append(tObj)
                            numBool = True
                            iID = iID + 1
            else:
                pass

    return tList

def ckNum(value):
    try:
        float(str(value))
        #print("True")
        return True
    except:
        #print("False")
        return False
    
def saveUnitData(iList):
    pathString = pickFile("xlsm")
    wb = load_workbook(pathString, keep_vba=True, read_only=False)
    ws = wb["Source Data"]
    i = 0 
    while True:
        vl= ws.cell(row = 3+i, column = 1).value
        if vl == None:
            break
        else:
            i = i + 1
    sRow = i + 3
    i = 0
    
    for item in iList:
        ws.cell(row = sRow+i, column = 1).value = 1 + ws.cell(row = sRow-1+i, column = 1).value
        ws.cell(row = sRow+i, column = 2).value = item.date
        ws.cell(row = sRow+i, column = 3).value = item.project
        ws.cell(row = sRow+i, column = 5).value = item.iDecp
        ws.cell(row = sRow+i, column = 9).value = item.contractor
        ws.cell(row = sRow+i, column = 11).value = item.unitRate
        ws.cell(row = sRow+i, column = 12).value = item.unit
        ws.cell(row = sRow+i, column = 13).value = item.qu
        i = i + 1
        
    wb.save(pathString)


def callDialogs():
    global unitPath
    global sectionPath

    if os.path.isfile("load_Paths.ini"):
        while True:
            print ("Do you want to use the existing loaded files (Y/N)? ")
            bExist = input(">>> ")
            if bExist == "Y" or bExist == "N":
                break
            else:
                print ("Invalid Entry")
    else:
        bExist = "N"

    if bExist == "N":
        print("Pick Unit Price File...")
        sleep(5)      
        pathString = pickFile("xlsm")

        print("Pick Section File...")
        sleep(5)      
        pathString2 = pickFile("xlsx")

        writePaths(pathString,pathString2)

        unitPath = pathString
        sectionPath = pathString2
        
    else:
        tArray = readPaths()
        unitPath = tArray[0]
        sectionPath = tArray[1]
        
def readPaths():
    fPaths = ConfigParser()
    fPaths.read("load_Paths.ini")

    path1 = fPaths['DEFAULT']['Unit Table Path']
    path2 = fPaths['DEFAULT']['Section List Path']
    
    return [path1,path2]

def writePaths(path1,path2):
    fPaths = ConfigParser()

    fPaths['DEFAULT'] = {'Unit Table Path' : path1,
                         'Section List Path':path2}
    with open("load_Paths.ini", 'w') as configfile:
        fPaths.write(configfile)

def compile_iObj():
    global unitPath
    pathString = unitPath
    wb = load_workbook(pathString)
    ws = wb["Source Data"]

    iList = []
    i = 4

    while True:
        if ws.cell(row = i, column = 1).value != None:
            tObj=itemObj()
            tObj.iNum  = ws.cell(row = i, column = 7).value      #Section Number
            tObj.iName = ws.cell(row = i, column = 4).value      #Name of Item
            tObj.iDecp = ws.cell(row = i, column = 5).value      #Description
            tObj.unit = ws.cell(row = i, column = 12).value       #Unit 
            tObj.unitRate = ws.cell(row = i, column = 11).value   #Unit Rate
            tObj.contractor = ws.cell(row = i, column = 9).value #Contractor 
            tObj.date = ws.cell(row = i, column = 2).value       #Date
            tObj.project = ws.cell(row = i, column = 3).value    #Project Name
            tObj.qu = ws.cell(row = i, column = 13).value         #Quantity
            tObj.sLife = ws.cell(row = i, column = 10).value      #Service Life
            tObj.cCode =ws.cell(row = i, column = 8).value       #Company Code
            tObj.notes = ws.cell(row = i, column = 14).value      #Notes
            tObj.gID = ws.cell(row = i, column = 15).value        #Group ID
            tObj.sID = ws.cell(row = i, column = 16).value        #SubGroup ID
            tObj.uni = ws.cell(row = i, column = 6).value         #Uniformat
            tObj.nID = ws.cell(row = i, column = 1).value         #Internal ID            
            iList.append(tObj)
            i = i + 1
        else:
            break

    return iList

def compile_gObj():
    global unitPath
    pathString = unitPath
    wb = load_workbook(pathString)
    ws = wb["Source Item List"]

    gList = []
    i = 3

    while True:
        if ws.cell(row = i, column = 1).value != None:
            tObj=itemObj()
            tObj.iNum  = ws.cell(row = i, column = 6).value      #Section Number
            tObj.iName = ws.cell(row = i, column = 3).value      #Name of Item
            tObj.iDecp = ws.cell(row = i, column = 4).value      #Description
            tObj.unit = ws.cell(row = i, column = 10).value       #Unit
            tObj.unitRate = "=AVGCOND(A*,B+)"                    #Unit Rate
            tObj.uniForm = ws.cell(row = i, column = 5).value    #Contractor 
            tObj.sLife = ws.cell(row = i, column = 8).value      #Service Life
            tObj.cCode =ws.cell(row = i, column = 7).value       #Company Code
            tObj.notes = ws.cell(row = i, column = 11).value      #Notes
            tObj.gID = ws.cell(row = i, column = 1).value        #Group ID
            tObj.sID = ws.cell(row = i, column = 2).value        #SubGroup ID
            gList.append(tObj)
            i = i + 1
        else:
            break

    return gList

def compile_sObj():
    global sectionPath
    pathString = sectionPath
    wb = load_workbook(pathString)
    ws = wb["Sections"]

    sList = []
    i = 2

    while True:
        if ws.cell(row = i, column = 1).value != None:
            tObj = sObj()
            tObj.sID = ws.cell(row = i, column = 1).value             #ID
            tObj.sNum  = ws.cell(row = i, column = 3).value           #Section Number
            tObj.sName = ws.cell(row = i, column = 2).value           #Section Name
            if ws.cell(row = i, column = 5).value != None:
                tObj.related = str(ws.cell(row = i, column = 5).value).split(",")         #Related
            else:
                tObj.related = None
            tObj.division = ws.cell(row = i, column = 7).value        #Section Division
            sList.append(tObj)
            i = i + 1
        else:
            break
    return sList

def compile_Cat():
    global unitPath
    pathString = unitPath
    wb = load_workbook(pathString, keep_vba=True, read_only=False)
    ws = wb["Categories"]

    tList = []
    i = 3

    while True:
        if ws.cell(row = i, column = 1).value != None:
            tObj = cObj()
            tObj.cID = ws.cell(row = i, column = 1).value
            tObj.cName = ws.cell(row = i, column = 2).value
            tList.append(tObj)
            i = i + 1
        else:
            break

    return tList

def savModUnit(iList):
    global unitPath
    pathString = unitPath
    wb = load_workbook(pathString, keep_vba=True, read_only=False)
    ws = wb["Source Data"]

    i = 4

    for item in iList:
            ws.cell(row = i, column = 7).value = item.iNum      #Section Number
            ws.cell(row = i, column = 4).value = item.iName      #Name of Item
            ws.cell(row = i, column = 5).value = item.iDecp     #Description
            ws.cell(row = i, column = 12).value = item.unit       #Unit 
            ws.cell(row = i, column = 9).value = item.contractor #Contractor 
            ws.cell(row = i, column = 2).value = item.date       #Date
            ws.cell(row = i, column = 3).value = item.project    #Project Name
            ws.cell(row = i, column = 13).value = item.qu         #Quantity
            ws.cell(row = i, column = 10).value = item.sLife 
            ws.cell(row = i, column = 8).value = item.cCode       #Company Code
            ws.cell(row = i, column = 14).value = item.notes      #Notes
            ws.cell(row = i, column = 15).value = item.gID        #Group ID
            ws.cell(row = i, column = 16).value = item.sID       #SubGroup ID
            ws.cell(row = i, column = 6).value = item.uni         #Uniformat
            i = i + 1

    wb.save(pathString)

def savMSingle(index, iList):
    global unitPath
    pathString = unitPath
    wb = load_workbook(pathString, keep_vba=True, read_only=False)
    ws = wb["Source Data"]

    item = iList[index]
    
    i = 4
    while True:
        vl= ws.cell(row = i, column = 1).value
        if vl == item.nID:
            break
        else:
            i = i + 1

    ws.cell(row = i, column = 7).value = item.iNum      #Section Number
    ws.cell(row = i, column = 4).value = item.iName      #Name of Item
    ws.cell(row = i, column = 5).value = item.iDecp     #Description
    ws.cell(row = i, column = 12).value = item.unit       #Unit 
    ws.cell(row = i, column = 9).value = item.contractor #Contractor 
    ws.cell(row = i, column = 2).value = item.date       #Date
    ws.cell(row = i, column = 3).value = item.project    #Project Name
    ws.cell(row = i, column = 13).value = item.qu         #Quantity
    ws.cell(row = i, column = 10).value = item.sLife 
    ws.cell(row = i, column = 8).value = item.cCode       #Company Code
    ws.cell(row = i, column = 14).value = item.notes      #Notes
    ws.cell(row = i, column = 15).value = item.gID        #Group ID
    ws.cell(row = i, column = 16).value = item.sID       #SubGroup ID
    ws.cell(row = i, column = 6).value = item.uni         #Uniformat

    wb.save(pathString)
    print("Saving entry to file....Done!")
    print("--------------------------------\n")

def savgObj(item):
    global unitPath
    pathString = unitPath
    wb = load_workbook(pathString, keep_vba=True, read_only=False)
    ws = wb["Source Item List"]

    i = 1 
    while True:
        vl= ws.cell(row = i, column = 1).value
        if vl == None:
            break
        else:
            i = i + 1

    ws.cell(row = i, column = 6).value = item.iNum      #Section Number
    ws.cell(row = i, column = 3).value = item.iName      #Name of Item
    ws.cell(row = i, column = 4).value = item.iDecp     #Description
    ws.cell(row = i, column = 10).value = item.unit     #Unit 
    ws.cell(row = i, column = 8).value = item.sLife 
    ws.cell(row = i, column = 7).value = item.cCode       #Company Code
    ws.cell(row = i, column = 11).value = item.notes      #Notes
    ws.cell(row = i, column = 1).value = item.gID        #Group ID
    ws.cell(row = i, column = 2).value = item.sID       #SubGroup ID
    ws.cell(row = i, column = 5).value = item.uniForm    #Uniformat

    wb.save(pathString)
   
def sMaker():
    print("Compiling new section list...")
    saveSectionData(getSectionList())
    
def uMaker():
    print("Gathering Bids for Unit Cost Tables")
    saveUnitData(gatherBids())
    
def mdMaker():
    print("Gathering Bids for Machine Learning")
    saveLite(gatherBids())
    
def uiChoice():
    pass

##def pUnit():
##    print("Preparing script, for further processing of the bids...")
##    wb = load_workbook(pickFile("xlsm"))
##    ws = wb["Source Data"]
##    i = 3
##
##    while True:
##        
##
##    wb.save("bidDataProcessed.xlsx")
