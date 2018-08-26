from openpyxl import load_workbook, Workbook
from tkinter import Tk
from tkinter import filedialog
from pathlib import Path
import os
from shutil import copy2
import numpy as np
import win32com.client as win32
from tClasses import sObj

#Word Application Start
word = win32.gencache.EnsureDispatch("Word.Application")
word.Visible = 0
word.DisplayAlerts = 0

#Constants
bidName = "00 54 22 - Schedule of Contract Unit Prices.docx"
sheetNames = ["Project Parameter","BidForm","SList","SectionList","Submittals"]
divisions = {0,1,2,3,4,5,6,7,8,9,15,23,26,32}
dirAuto = os.path.dirname(os.path.realpath(__file__))

#Open Dialog Function
def openDialog(fExt):
    oRoot = Tk()
    fname = filedialog.askopenfilename(initialdir = "C:/",
                                       title = "Select File",
                                       filetypes = ((fExt+" files","*."+fExt),
                                                    ("All File Types","*.*"))
                                       )
    oRoot.destroy()
    if fname == "":
        return ""
    else:
        return fname

#Read Info from it
def specPara(fname):
    wb = load_workbook(fname)
    ws = wb[sheetNames[0]]
    pValues = []
    keyValues = []
    i = 2
    
    while True:
        if ws["A"+str(i)].value != None:
            pValues.append([ws["B"+str(i)].value])
            keyValues.append([ws["C"+str(i)].value])
            i = i + 1
        else:
            break
        
    print("Parameters Loaded")
    return pValues, keyValues

def loadSList(fname):
    secList = []
    wb = load_workbook(fname)
    ws = wb[sheetNames[3]]
    i = 2
    
    while True:
        tID = ws["A"+str(i)].value
        if tID != None:
            section = sObj()
            section.sID = ws["A"+str(i)].value
            section.sNum = ws["B"+str(i)].value
            section.sPage = ws["C"+str(i)].value
            section.sName = ws["D"+str(i)].value
            section.state = ws["E"+str(i)].value
            if ws["F"+str(i)].value != "" and ws["F"+str(i)].value != None:
                section.related = (ws["F"+str(i)].value).split(",")
            secList.append(section)
            i = i + 1
            print("Section Loaded")
        else:
            break
    print ("Section List Loaded")
    return secList
    
    
#Get Documents from each division
def coDivi(dNum,secList):
    if not os.path.exists(dirAuto+"\DIVISION " + str(dNum)):
        print ("Division folder not present in home directory")
        return None
    
    if not os.path.exists(dirAuto + "\Output Spec"):
        os.makedirs(dirAuto + "\Output Spec")

    for f in Path(dirAuto+"\DIVISION " + str(dNum)).iterdir():
        tString2 = dirAuto + "\DIVISION " + str(dNum) + "\\"
        tString = str(f).replace(tString2,"")
        for s in secList:
            if tString.find(str(s.sNum)) is not -1:
                if s.state == 1:
                    copy2(str(f),dirAuto + "\Output Spec" +"\\" + tString)
    print ("Division " + str(dNum) + " Compiled")

def compilerSpec(fName):
    pValue, kValue = specPara(fName)
    secList = loadSList(fName) 
    for num in divisions:
        print("Division " + str(num)+" Start...")
        coDivi(num,secList)
    paraDivi(pValue,kValue)

def paraDivi(pValues,keyValues):
    for f in Path(dirAuto + "\Output Spec").iterdir():
        if(str(f).find("~$")==-1):
            print("Processing " + str(f) + "...")
            tdoc = word.Documents.Open(str(f))
            for i in range(len(pValues)):
                kV = str(keyValues[i][0])
                docx_replace_regex(tdoc.Content,kV,str(pValues[i][0]))
                docx_replace_regex2(tdoc,kV,str(pValues[i][0]))
                docx_replace_regex3(tdoc,kV,str(pValues[i][0]))
            tdoc.SaveAs(str(f))
            tdoc.Close()
        
    print ("Operation Complete")
        
#Replace Functions for Docx
#Core 
def docx_replace_regex(doc_obj, regex , replace):
        doc_obj.Find.Text = regex
        doc_obj.Find.Replacement.Text = replace
        doc_obj.Find.Execute(Replace=2, Forward=True, Wrap=0) 

#Headers & Footers
def docx_replace_regex2(doc_obj, regex , replace):
    for s in doc_obj.Sections:
        hColl = s.Headers
        fColl = s.Footers
        for j in hColl:
            docx_replace_regex(j.Range,regex,replace)
        for j in fColl:
            docx_replace_regex(j.Range,regex,replace)

#Tables
def docx_replace_regex3(doc_obj, regex , replace):
    for t in doc_obj.Tables:
        for row in t.Rows:
            for cell in row.Cells:
                docx_replace_regex(cell.Range,regex,replace)
    
#Main Loop
print("Running Spec Automation Script...")
fname = openDialog("xlsx")
if fname == "":
   word.Quit()
   exit()
compilerSpec(fname)
word.Quit()
